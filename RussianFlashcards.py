from tkinter import *
from random import randint
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

root = Tk()
root.title('Russian Language Flashcard App')
root.geometry("700x520")

# load existing spreadsheet
try:
    wb = load_workbook('RussianTranslations.xlsx')
except:
    print("Make sure RussianTranslations.xlsx is in the same folder")

# Global variables
random_word = 0

# Functions
def clear():
    # Clear the screen
    russian_word.config(text="")
    answer_label.config(text="")
    result_label.config(text="")
    hint_label.config(text="")
    my_entry.delete(0, END)
    next_button.config(state=DISABLED)
    
def load_mode(mode):
    clear()
    wb.active = wb[mode]
    global ws
    ws = wb.active
    next()
    my_entry.focus()
    input_button.config(state=NORMAL)
    hint_button.config(state=NORMAL)
    show_answer_button.config(state=NORMAL)
    

def next():
    # Create random selection and store last word to prevent repeats
    global random_word
    last_word = random_word
    # Clear the screen
    clear()
    # find a new card index, different than the previous
    while (random_word == last_word):
        random_word = randint(2, ws.max_row)
    # update label with Russian Word
    russian_word.config(text=ws[random_word][0].value)
    root.unbind("<Return>")
    root.bind("<Return>", return_result)
    my_entry.focus()
    
def determine_input():
    if my_entry.get().lower() in (ws[random_word][2].value, ws[random_word][3].value, ws[random_word][4].value, ws[random_word][5].value):
        result_label.config(text="Correct!")
        next_button.config(state=NORMAL)
        root.unbind("<Return>")
        root.bind("<Return>", return_next)
    else: 
        result_label.config(text="Incorrect")

def return_result(event):
    determine_input()

def return_next(event):
    next()

def show_answer():
    answer_label.config(text=ws[random_word][2].value)
    my_entry.focus()
        
def hint():
    hint_label.config(text=ws[random_word][1].value)
    my_entry.focus()

# Mode Buttons
mode_button_frame = Frame(root)
mode_button_frame.pack(pady=20)

alphabet_button = Button(mode_button_frame, text="Alphabet", command=lambda:load_mode("Alphabet"))
alphabet_button.grid(row=0, column=0, padx=20)

nouns_button = Button(mode_button_frame, text="Nouns", command=lambda:load_mode("Nouns"))
nouns_button.grid(row=0, column=1,padx=20)

verbs_button = Button(mode_button_frame, text="Verbs", command=lambda:load_mode("Verbs"))
verbs_button.grid(row=0, column=2, padx=20)

common100_button = Button(mode_button_frame, text="Common Words", command=lambda:load_mode("Common100"))
common100_button.grid(row=0, column=3, padx=0)

onetoten_button = Button(mode_button_frame, text="1-10", command=lambda:load_mode("1-10"))
onetoten_button.grid(row=0, column=4, padx=20)

upto100_button = Button(mode_button_frame, text="11-100", command=lambda:load_mode("11-100"))
upto100_button.grid(row=0, column=5, padx=20)

ordinal_button = Button(mode_button_frame, text="Ordinal", command=lambda:load_mode("Ordinal"))
ordinal_button.grid(row=0, column=6, padx=20)


# Flashcard Labels
russian_word = Label(root, text="Select a mode above", font=("Helvetica", 36))
russian_word.pack(pady=(50,20))

hint_label = Label(root, text="", font=("Helvetica", 14))
hint_label.pack(pady=10)

answer_label = Label(root, text="", font=("Helvetica", 30))
answer_label.pack(pady=(20, 30))


# Input Frame
input_frame = Frame(root)
input_frame.pack(pady=10)

my_entry = Entry(input_frame, font=("Helvetica", 18))
my_entry.pack(side=LEFT, padx=10)

input_button = Button(input_frame, text="Input", command=determine_input, state=DISABLED)
input_button.pack(side=RIGHT, padx=10)

result_label = Label(root, text="", font=("Helvetica", 16))
result_label.pack(pady=10)


# Action Buttons
action_button_frame = Frame(root)
action_button_frame.pack(pady=20)

hint_button = Button(action_button_frame, text="Hint", command=hint, state=DISABLED)
hint_button.grid(row=0, column=0, padx=20)

show_answer_button = Button(action_button_frame, text="Show Answer", command=show_answer, state=DISABLED)
show_answer_button.grid(row=0, column=1, padx=0)

next_button = Button(action_button_frame, text="Next", command=next, state=DISABLED)
next_button.grid(row=0, column=2,padx=20)


# Run program
root.mainloop()